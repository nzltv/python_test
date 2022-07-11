import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import time
from selenium.webdriver.common.by import By
import datetime
import calendar
import pandas as pd
from openpyxl.utils import get_column_letter
import SendEmail

def main():
    months = ['', 'Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября', 'Декабря']
    #получаем первый день месяца
    firstDay='1'
    #получаем последний день месяца
    lastDay = datetime.datetime.now().replace(day = calendar.monthrange(datetime.datetime.now().year, datetime.datetime.now().month)[1]).strftime("%d")
    #получаем текущий месяц
    currentMonth = months[datetime.datetime.now().month]
    #получаем текущий год
    currentYear = '2021' #получить текущий год - datetime.datetime.now().strftime("%Y")

    driver = webdriver.Chrome()
    # разворачиваем окно во весь экран
    driver.maximize_window()

    #Открыть сайт https://www.moex.com.
    driver.get("https://www.moex.com/")

    #Нажать на кнопку Меню
    menuBtn = driver.find_element_by_xpath("//*[contains(@class, 'header-menu')]/span[1]/button")                                
    menuBtn.click()
    time.sleep(3) 

    #выбрать Срочный рынок
    marketBtn = driver.find_elements_by_xpath("//a[contains(text(), 'Срочный рынок')]")[1]
    marketBtn.click()
    time.sleep(3) 

    #если есть кнопка согласен - нажимаем её
    if check_exists_by_xpath(driver, "//a[contains(text(), 'Согласен')]"):
        driver.find_element_by_xpath("//a[contains(text(), 'Согласен')]").click()

    #выбрать Индикативные курсы
    time.sleep(3) 
    coursesBtn = driver.find_element_by_xpath("//a[contains(text(), 'Индикативные курсы')]")
    coursesBtn.click()
    # выбираем валюту 
    driver.find_element_by_id('ctl00_PageContent_CurrencySelect').send_keys('USD/RUB')
    # выбираем даты
    selectDates(driver, firstDay, lastDay, currentMonth, currentYear)
    # если записи есть - формируем таблицу
    if check_exists_by_xpath(driver, "//b[contains(text(), 'Записей не найдено')]"):
        FirstTable = pd.DataFrame()
    else:
        FirstTable = getTable(driver,"//table[@class= 'tablels']", 'USD')

    # выбираем валюту
    driver.find_element_by_id('ctl00_PageContent_CurrencySelect').send_keys('EUR/RUB')
    selectDates(driver, firstDay, lastDay, currentMonth, currentYear)
    # если записи есть - формируем таблицу
    if check_exists_by_xpath(driver, "//b[contains(text(), 'Записей не найдено')]"):
        SecondTable = pd.DataFrame()
    else:
        SecondTable = getTable(driver,"//table[@class= 'tablels']", 'EUR')
    
    # если таблицы не пустые - объединяем
    if (FirstTable.shape[0] != 0 and SecondTable.shape[0] != 0):
        resultTable = pd.merge(FirstTable, SecondTable, left_on='USD_Дата', right_on='EUR_Дата', how='left').drop('EUR_Дата', axis=1)
        resultTable['Изменение'] = 0
        for index, row in resultTable.iterrows():
            resultTable['Изменение'][index] = resultTable['EUR_Значение курса основного клиринга'][index] / resultTable['USD_Значение курса основного клиринга'][index]

        resultStr = 'Получена информация по USD и EUR'
    elif FirstTable.shape[0] != 0:
        resultTable = FirstTable
        resultStr = 'Получена информация только по USD'
    elif SecondTable.shape[0] != 0:
        resultTable = SecondTable
        resultStr = 'Получена информация только по EUR'
    # вносим правки в Excel и отправляем письмо
    changeExcelAndSendMail(resultTable, resultStr)
    print('done')

def changeExcelAndSendMail(resultTable, resultStr):
    # считываем Excel с результатами
    writer = pd.ExcelWriter("result.xlsx", engine='xlsxwriter')
    # записываем результат работы
    resultTable.to_excel(writer,index=0, sheet_name="result")
    # сохраняем Excel
    writer.save()
    # получаем объект книги Excel
    wb = openpyxl.load_workbook("result.xlsx")
    # получаем лист
    ws = wb["result"]
    # запоминаем число строк для отправки в письме
    rowsCount = len(resultTable)
    # меняем формат на финансовый
    for row in range(2, ws.max_row+1):
        ws["{}{}".format("B", row)].number_format = '#,##0.00₽'
        ws["{}{}".format("D", row)].number_format = '#,##0.00₽'
        ws["{}{}".format("F", row)].number_format = '#,##0.00₽'
        ws["{}{}".format("H", row)].number_format = '#,##0.00₽'
        ws["{}{}".format("J", row)].number_format = '#,##0.00₽'
    # изменяем ширину столбца (ориентируясь на максимальное кол-во символов в столбце)
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.1
    # сохраняем Excel
    wb.save("result.xlsx")
    # оправляем письмо
    SendEmail.SendEmail("result.xlsx", rowsCount, resultStr)

def selectDates(driver, firstDay, lastDay, currentMonth, currentYear):
    #выбираем первое число месяца
    driver.find_element_by_id('d1day').send_keys(firstDay)
    driver.find_element_by_id('d1month').send_keys(currentMonth)
    driver.find_element_by_id('d1year').send_keys(currentYear)
    #выбираем последнее число месяца
    driver.find_element_by_id('d2day').send_keys(lastDay)
    driver.find_element_by_id('d2month').send_keys(currentMonth)
    driver.find_element_by_id('d2year').send_keys(currentYear)
    time.sleep(3)
    #Нажать на кнопку Показать
    submitBtn = driver.find_element_by_name('bSubmit')
    driver.execute_script("arguments[0].click();", submitBtn)
    time.sleep(3)

def getTable(driver, xpath, currency):
    
    # формируем словарь с заголовками таблицы
    headres = [currency + '_Дата' , currency + '_Значение курса промежуточного клиринга', currency + '_Время ', currency + '_Значение курса основного клиринга', currency + '_Время']
    # создаем таблицу
    df_result = pd.DataFrame(columns=headres)
    # получаем таблицу курсов с сайта
    tbl = driver.find_elements_by_xpath (xpath)[0]
    # перебираем строки таблицы с сайта(кроме заголовков)
    for r in range(2, len(tbl.find_elements(By.TAG_NAME, "tr"))):
        row = tbl.find_elements(By.TAG_NAME, "tr")[r]
        # создаем словарь для добавления значений в итоговую таблицу
        dict = {headres[0]: '',
                headres[1]: 0,
                headres[2]: '',
                headres[3]: 0,
                headres[4]: ''}
        # помещаем значения из строки сайта в словарь
        dict[headres[0]] = row.find_elements(By.TAG_NAME, "td")[0].text
        dict[headres[1]] = float(row.find_elements(By.TAG_NAME, "td")[1].text.replace(',','.'))
        dict[headres[2]] = row.find_elements(By.TAG_NAME, "td")[2].text
        dict[headres[3]] = float(row.find_elements(By.TAG_NAME, "td")[3].text.replace(',','.'))
        dict[headres[4]] = row.find_elements(By.TAG_NAME, "td")[4].text
        # добавления строки в итоговую таблицу
        df_result= df_result.append(dict, ignore_index = True)
        # очищаем словарь
        dict.clear()
    return df_result
    
def check_exists_by_xpath(driver, xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        SendEmail.SendEmailError(str(e.args[0]))

    
